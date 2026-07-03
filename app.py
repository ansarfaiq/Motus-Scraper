import re
import sys
import os
import csv
import threading
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
import streamlit as st

# پلے رائٹ کے براؤزرز اور ان کی سسٹم ڈیپینڈینسیز کو کلاؤڈ پر خود بخود انسٹال کرنے کے لیے
@st.cache_resource
def install_playwright_browsers():
    # یہ لائن کرومیم براؤزر اور اس کے لیے ضروری تمام لینکس فائلیں ایک ہی بار ڈاؤن لوڈ کر دے گی
    os.system("playwright install --with-deps chromium")

# فنکشن کو صرف ایک بار کال کریں
install_playwright_browsers()

# ایکسل پڑھنے کے لیے
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# =====================================================
# 1. SCRAPER کلاس (تمام فنکشنز اور لاجک بالکل سیم ہیں)
# =====================================================
class Scraper:
    def __init__(self, log_callback=None, row_callback=None):
        self.log_callback = log_callback
        self.row_callback = row_callback
        self.results = []

    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)

    def clean_extracted_email(self, raw_email):
        if not raw_email or raw_email == "Not Available":
            return "Not Available"
        cleaned = raw_email.strip()
        if cleaned.lower().startswith("businessemail"):
            cleaned = cleaned[13:]
        elif cleaned.lower().startswith("email"):
            cleaned = cleaned[5:]
        if cleaned.lower().endswith("company"):
            cleaned = cleaned[:-7]
        return cleaned.strip()

    def scrape_single_dot(self, dot_num, page):
        data = {
            "USDOT": dot_num,
            "Legal_Name": "Not Available",
            "Email": "Not Available",
            "Phone": "Not Available",
            "State": "Not Available",
            "Status": "Not Available",
        }
        try:
            page.goto(f"https://motus.dot.gov/customer/{dot_num}/account", timeout=30000)
            page.wait_for_timeout(4500)
            visible_text = page.locator("body").inner_text()

            # Name
            name_match = re.search(r'Business Name\s+([A-Za-z0-9\s\.\&\-]+?)(?:\s+Duns|\s+Doing|\s+Principal|\s+Form of Business|\s+$|\n)', visible_text, re.IGNORECASE)
            if name_match:
                potential_name = name_match.group(1).strip()
                if potential_name and len(potential_name) > 2:
                    if not any(x in potential_name.lower() for x in ["united states", "official", "website", "www", "duns", "form of business", "here's", "how you know"]):
                        data["Legal_Name"] = potential_name

            if data["Legal_Name"] == "Not Available":
                lines = visible_text.split('\n')
                for i, line in enumerate(lines):
                    if "Business Name" in line:
                        if i+1 < len(lines):
                            next_line = lines[i+1].strip()
                            if next_line and len(next_line) > 2:
                                if not any(x in next_line.lower() for x in ["duns", "form of business", "here's", "how you know", "official", "website"]):
                                    data["Legal_Name"] = next_line
                                    break

            # Email
            email_match = re.search(r'Business Email\s+([\w\.-]+@[\w\.-]+\.\w+)', visible_text, re.IGNORECASE)
            if email_match:
                data["Email"] = self.clean_extracted_email(email_match.group(1))
            if data["Email"] == "Not Available":
                emails = re.findall(r"[\w\.-]+@[\w\.-]+\.\w+", visible_text)
                if emails:
                    clean_emails = [e for e in emails if not any(x in e.lower() for x in ["bootstrap", "w3.org", "png", "schema", "example", "test", "@"])]
                    if clean_emails:
                        data["Email"] = self.clean_extracted_email(clean_emails[0])

            # Phone
            phone_match = re.search(r'Business Telephone No\.\s+([\+\d\s\(\)\-]+)', visible_text, re.IGNORECASE)
            if phone_match:
                phone_raw = phone_match.group(1).strip()
                phone_clean = re.search(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', phone_raw)
                if phone_clean:
                    data["Phone"] = phone_clean.group(0)

            # State
            state_match = re.search(r'\b([A-Z]{2})\s+\d{5}\b', visible_text)
            if state_match:
                data["State"] = state_match.group(1)

            # Status
            if "active" in visible_text.lower() or "authorized" in visible_text.lower():
                data["Status"] = "Active / Authorized"
            else:
                data["Status"] = "Inactive"

        except Exception as e:
            self.log(f"⚠️ Error on DOT {dot_num}: {str(e)}")

        return data

    def run(self, input_data):
        self.results = []
        self.log("🚀 Scraper is starting...")

        dot_numbers = []
        for item in input_data:
            found = re.findall(r'\b\d+\b', str(item))
            dot_numbers.extend(found)

        if not dot_numbers:
            self.log("❌ No valid USDOT numbers found.")
            return []

        self.log(f"✅ Found {len(dot_numbers)} USDOT numbers.")

        with sync_playwright() as p:
            try:
                browser = p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-dev-shm-usage'])
            except Exception as e:
                self.log(f"❌ Browser launch failed: {e}")
                return []

            page = browser.new_page()
            total = len(dot_numbers)

            for idx, dot in enumerate(dot_numbers, 1):
                self.log(f"⏳ Processing {idx}/{total}: DOT {dot}")
                data = self.scrape_single_dot(dot, page)

                if data["Legal_Name"] != "Not Available" and len(data["Legal_Name"]) > 2:
                    if not any(x in data["Legal_Name"].lower() for x in ["official website", "united states", "www", "business name", "here's", "how you know"]):
                        self.results.append(data)
                        self.log(f"✅ Found: {data['Legal_Name']} | Email: {data['Email']}")
                        
                        if self.row_callback:
                            self.row_callback(data)
                    else:
                        self.log(f"⏭️ Skipped DOT {dot} (junk name)")
                else:
                    self.log(f"⏭️ Skipped DOT {dot} (no valid Business Name)")

            browser.close()

        self.log(f"✅ Scraping complete! Total {len(self.results)} records.")
        return self.results


# =====================================================
# 2. STREAMLIT WEB UI
# =====================================================
st.set_page_config(page_title="Motus Data Scraper", page_icon="🚛", layout="wide")

st.title("🚛 Motus Data Scraper")
st.subheader("⚡ Real-time Table View | Upload TXT / CSV / XLSX ⚡")

# سیشن اسٹیٹ ڈیٹا کو سنبھالنے کے لیے
if "scraped_data" not in st.session_state:
    st.session_state.scraped_data = []
if "log_messages" not in st.session_state:
    st.session_state.log_messages = "✅ Ready | Upload a file or type DOT numbers to begin."

# ان پٹ کے دو طریقے
col1, col2 = st.columns([2, 1])

with col1:
    raw_text = st.text_area("📝 Enter USDOT Numbers manually (comma or space separated):", height=100)

with col2:
    uploaded_file = st.file_uploader("📂 Load File (.txt, .csv, .xlsx)", type=["txt", "csv", "xlsx"])

# فائل سے ڈیٹا نکالنے کا لاجک
input_dots = []
if uploaded_file:
    ext = uploaded_file.name.split(".")[-1].lower()
    content = ""
    if ext == "txt":
        content = uploaded_file.read().decode("utf-8")
    elif ext == "csv":
        content = uploaded_file.read().decode("utf-8")
    elif ext == "xlsx" and load_workbook:
        wb = load_workbook(uploaded_file, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                content += " " + " ".join([str(c) for c in row if c])
    
    input_dots = re.findall(r'\b\d+\b', content)
    if input_dots:
        st.success(f"Loaded {len(set(input_dots))} numbers from file.")

if raw_text:
    manual_dots = re.findall(r'\b\d+\b', raw_text)
    input_dots.extend(manual_dots)

# یونیک نمبرز نکالنا
final_dots = sorted(list(set([d for d in input_dots if d.isdigit()])))

# لائیو لاگز اور اسٹیٹس کے لیے کنٹینر
status_box = st.info(st.session_state.log_messages)

def update_live_ui(msg):
    st.session_state.log_messages = msg
    status_box.info(msg)

def live_row_added(row_data):
    st.session_state.scraped_data.append(row_data)

# بٹنز کنٹرولز
btn_col1, btn_col2 = st.columns([1, 5])

with btn_col1:
    if st.button("▶️ Start Scraping", type="primary"):
        if not final_dots:
            st.warning("Please enter or upload USDOT numbers first!")
        else:
            st.session_state.scraped_data = [] # ٹیبل کلیئر کریں
            scraper = Scraper(log_callback=update_live_ui, row_callback=live_row_added)
            
            with st.spinner("Scraping data... Please wait."):
                scraper.run(final_dots)
            st.success("Scraping complete!")

with btn_col2:
    if st.button("🗑️ Clear Table"):
        st.session_state.scraped_data = []
        st.session_state.log_messages = "🗑️ Table cleared."
        st.rerun()

# لائیو ڈیٹا ٹیبل ڈسپلے
st.write("### 📊 Extracted Data:")
if st.session_state.scraped_data:
    st.dataframe(st.session_state.scraped_data, use_container_width=True)
    
    # CSV ڈاؤن لوڈ بٹن
    # ڈیٹا کو سٹرکچرڈ فارمیٹ میں بدلیں
    csv_rows = []
    for item in st.session_state.scraped_data:
        csv_rows.append(f"{item['USDOT']},{item['Legal_Name']},{item['Email']},{item['Phone']},{item['State']},{item['Status']}\n")
    
    csv_data = "USDOT,Legal_Name,Email,Phone,State,Status\n" + "".join(csv_rows)
    
    st.download_button(
        label="💾 Save CSV",
        data=csv_data,
        file_name="scraped_carriers.csv",
        mime="text/csv"
    )
else:
    st.variant = st.write("No data available in table.")
